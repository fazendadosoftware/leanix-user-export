import json
import requests
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import base64
import time
import os
import sys
import math

def getApiToken():
    with open('./lxr.json') as json_file:
        data = json.load(json_file)
        return data['apitoken']

def getHost():
    with open('./lxr.json') as json_file:
        data = json.load(json_file)
        return data['host']

def getAccessToken(api_token):
    url = mtm_base_url+"/oauth2/token"
    response = requests.post(url=url, auth=('apitoken', api_token), data={'grant_type': 'client_credentials'})
    response.raise_for_status()
    access_token = response.json()['access_token']
    return access_token

# Function to decipher the access_token
def getAccessTokenJson(access_token):
  payload_part = access_token.split('.')[1]
  # fix missing padding for this base64 encoded string.
  # If number of bytes is not dividable by 4, append '=' until it is.
  missing_padding = len(payload_part) % 4
  if missing_padding != 0:
    payload_part += '='* (4 - missing_padding)
  payload = json.loads(base64.b64decode(payload_part))
  return payload

def getWorkspaceId(access_token_json):
    return access_token_json['principal']['permission']['workspaceId']

def getWorkspaceName(access_token, workspaceId):
    url = mtm_base_url + '/workspaces/' + workspaceId
    headers = { 'Authorization': 'Bearer ' + access_token, 'Content-Type': 'application/json' }
    response = callGet(url, headers)
    return response['data']['name']

def getWorkspacePermissions(access_token, page = 1, size = 40):
  workspace_id = getWorkspaceId(getAccessTokenJson(access_token))
  url = mtm_base_url + "/workspaces/" + workspace_id + "/permissions"
  params = { 'page': page, 'size': size, 'status': '!ARCHIVED' }
  headers = { 'Authorization': 'Bearer ' + access_token, 'Content-Type': 'application/json' }
  print('Fetching permissions page 1...')
  response = callGet(url, headers, params)
  total = response['total']
  permissions = response['data']
  if page == 1:
    pages = math.ceil(total / size)
    for p in range(2, pages + 1):
      print('Fetching permissions page ' + str(p) + ' of ' + str(pages))
      params['page'] = p
      permissions = permissions + callGet(url, headers, params)['data']
  return permissions

# General function to call GraphQL given a query
def callGet(url, headers = {}, params = {}):
  response = requests.get(url=url, headers=headers, params = params)
  response.raise_for_status()
  return response.json()

def callPost(request_url, header, data):
    try:
        response = requests.post(
            url=request_url, headers=header, data=json.dumps(data))
        response.raise_for_status()
    except requests.exceptions.HTTPError as err:
        print(request_url)
        print(json.dumps(data))
        print(err)
        exit
    return response.json()

mtm_base_url = 'https://' + getHost() + '/services/mtm/v1'
pathfinder_base_url = 'https://' + getHost() + '/services/pathfinder/v1'

access_token = getAccessToken(getApiToken())
workspaceName = getWorkspaceName(access_token, getWorkspaceId(getAccessTokenJson(access_token)))

print("Logged in workspace " + workspaceName)

permissions = getWorkspacePermissions(access_token)

rows = [['User ID', 'Username', 'Email', 'First Name', 'Last Name', 'Permission ID', 'Status (Permission)', 'Role', 'Customer Roles', 'Access Control Entities']]
for permission in permissions:
  user = permission.get('user')
  user_id = user.get('id')
  username = user.get('userName')
  email = user.get('email')
  first_name = user.get('firstName')
  last_name = user.get('lastName')
  permission_id = permission.get('id')
  status = permission.get('status')
  role = permission.get('role')
  customer_roles = permission.get('customerRoles')
  access_control_entities = permission.get('accessControlEntities')
  rows.append([user_id, username, email, first_name, last_name, permission_id, status, role, customer_roles, access_control_entities])

wb = Workbook()
ws = wb.active


# Auto-adjust column width size...
column_widths = []
for row in rows:
  ws.append(row)
  for i, cell in enumerate(row):
    len_cell = 0
    if cell is not None:
      len_cell = len(cell)
    if len(column_widths) > i:
      if len_cell > column_widths[i]:
          column_widths[i] = len_cell
    else:
      column_widths += [len_cell]

for i, column_width in enumerate(column_widths):
    ws.column_dimensions[get_column_letter(i+1)].width = column_width

wb.save(workspaceName + '_permissions.xlsx')

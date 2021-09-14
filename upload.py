import json
import base64
import requests
import sys
from openpyxl import load_workbook

FILENAME = 'upload.xlsx'

# https://docs.python-requests.org/en/latest/user/advanced/#proxies
# uncomment lines below for setting up proxy operation
# os.environ["HTTP_PROXY"] = "http://127.0.0.1:1234"
# os.environ["HTTPS_PROXY"] = "https://127.0.0.1:1234"

def getHost():
  with open('./lxr.json') as json_file:
    data = json.load(json_file)
    return data['host']
      
def getApiToken():
  with open('./lxr.json') as json_file:
    data = json.load(json_file)
    return data['apitoken']

mtm_base_url = 'https://' + getHost() + '/services/mtm/v1'

def getAccessToken(api_token):
  url = mtm_base_url+"/oauth2/token"
  response = requests.post(url=url, auth=('apitoken', api_token), data={'grant_type': 'client_credentials'})
  response.raise_for_status()
  access_token = response.json()['access_token']
  return access_token

# Function to decipher the access_token
def getAccessTokenJson(access_token):
  payload_part = access_token.split('.')[1]
  # fix missing padding for this base64 encoded string.
  # If number of bytes is not dividable by 4, append '=' until it is.
  missing_padding = len(payload_part) % 4
  if missing_padding != 0:
    payload_part += '='* (4 - missing_padding)
  payload = json.loads(base64.b64decode(payload_part))
  return payload

def getWorkspaceId(access_token_json):
    return access_token_json['principal']['permission']['workspaceId']
  
def getWorkspaceName(access_token, workspaceId):
  url = mtm_base_url + '/workspaces/' + workspaceId
  headers = { 'Authorization': 'Bearer ' + access_token, 'Content-Type': 'application/json' }
  response = requests.get(url=url, headers=headers)
  response.raise_for_status()
  return response.json()['data']['name']

def fetch_permission_by_id(access_token, permission_id):
  workspace_id = getWorkspaceId(getAccessTokenJson(access_token))
  url = mtm_base_url + "/workspaces/" + workspace_id + "/permissions/" + permission_id
  headers = { 'Authorization': 'Bearer ' + access_token }
  response = requests.get(url=url, headers=headers)
  response.raise_for_status()
  permission = response.json()['data']
  return permission

def update_permission(access_token, permission):
  url = mtm_base_url + '/permissions'
  headers = {
    'Authorization': 'Bearer ' + access_token,
    'Content-Type': 'application/json'
  }
  response = requests.post(url=url, headers=headers, data=json.dumps(permission))
  try:
    response.raise_for_status()
  except:
    print('Exception: ' + str(response.status_code) + ' ' + response.text)
    raise
  data = response.json()['data']
  return data

def fetch_user_by_id(access_token, user_id):
  url = mtm_base_url + "/users/" + user_id
  headers = { 'Authorization': 'Bearer ' + access_token }
  response = requests.get(url=url, headers=headers)
  response.raise_for_status()
  user = response.json()['data']
  return user

def update_user(access_token, user):
  url = mtm_base_url + "/users/" + user_id
  headers = {
    'Authorization': 'Bearer ' + access_token,
    'Content-Type': 'application/json'
  }
  response = requests.put(url=url, headers=headers, data=json.dumps(user))
  try:
    response.raise_for_status()
  except:
    print('Exception: ' + str(response.status_code) + ' ' + response.text)
    raise
  data = response.json()['data']
  return data

access_token = getAccessToken(getApiToken())

workspaceName = getWorkspaceName(access_token, getWorkspaceId(getAccessTokenJson(access_token)))

print("Logged in workspace " + workspaceName)

wb = load_workbook(FILENAME)

# select first worksheet
ws = wb.active

# iterate rows
i = -1
for row in ws.values:
  i = i + 1
  if i == 0:
    continue
  user_id = row[0]
  username = row[1]
  email = row[2]
  first_name = row[3]
  last_name = row[4]
  permission_id = row[5]
  status = row[6]
  role = row[7]
  customer_roles = row[8]
  access_control_entities = row[9]
  permission = fetch_permission_by_id(access_token, permission_id)
  # check if user id matches
  if permission['user']['id'] != user_id:
    raise Exception('Invalid user id, expected ' + user_id, ', got ' + permission['user']['id'])
  if permission['user']['userName'] != username:
    user = fetch_user_by_id(access_token, user_id)
    user['userName'] = username
    update_user(access_token, user)
  permission['user']['userName'] = username
  permission['user']['email'] = email
  permission['user']['firstName'] = first_name
  permission['user']['lastName'] = last_name
  permission['status'] = status
  permission['role'] = role
  permission['customerRoles'] = customer_roles
  permission['accessControlEntities'] = access_control_entities
  print('Updated permission {id} {username} ({i}/{total})'.format(id=permission_id, username=username, i=i, total=ws.max_row - 1))
  update_permission(access_token, permission)
